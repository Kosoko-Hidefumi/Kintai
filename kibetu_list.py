import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import re
import json
from openai import OpenAI
from typing import Dict, Optional
from dotenv import load_dotenv

# .envファイルから環境変数を読み込む
load_dotenv()


def normalize_name(name):
    """
    名前を正規化する（括弧部分を削除）
    例: '筧咲陽子(渡部)' -> '筧咲陽子'
    """
    if pd.isna(name):
        return ''
    name_str = str(name)
    # 全角・半角の括弧とその中身を削除
    normalized = re.sub(r'[（(].+?[）)]', '', name_str)
    return normalized.strip()


# 沖縄県の施設リスト（正規化前）
OKINAWA_FACILITIES_RAW = [
    "県立宮古病院", "県立北部病院", "県立八重山病院", "県立中部病院",
    "伊平屋診療所", "伊是名診療所", "西表西部診療所", "小浜診療所",
    "南部医療センター附属北大東診療所", "八重山病院附属西表西武診療所",
    "八重山病院", "北部病院附属伊平屋診療所", "琉大病院　麻酔科",
    "琉大病院　消化器　外科", "宮古病院", "中部病院", "北部病院",
    "南部医療センター附属南大東診療所", "座間味診療所", "阿嘉診療所",
    "県立八重山病院付属大原診療所", "県立北部病院付属伊是名診療所",
    "県立北部病院所属伊平屋診療所", "精和病院", "日本赤十字センター",
    "県立南部医療センター所属粟国診療所", "琉大附属病院　整形外科",
    "県立南部医療センター所属南大東診療所", "琉大病院",
    "県立南部医療センター所属渡名喜診療所", "県立南部医療センター所属波照間診療所",
    "県立南部医療センター所属阿嘉診療所", "県立北部病院所属伊是名診療所",
    "粟国診療所", "北大東診療所", "大原診療所", "南部医療センター",
    "県立南部医療センター所属座間味診療所", "県立北部病院所属伊是名診療所",
    "琉大附属病院", "浦添総合病院", "琉球大学医学部附属病院　泌尿器科",
    # 追加：琉球大学関連の表記バリエーション
    "琉球大学医学部附属病院", "琉球大学病院", "琉大", "琉球大学附属病院",
    "琉球大学医学部付属病院", "琉大医学部附属病院"
]


def normalize_facility_name(facility_name: str, client: Optional[OpenAI] = None, cache: Optional[Dict[str, str]] = None) -> str:
    """
    OpenAI APIを使って施設名を正規化する
    
    Parameters:
    -----------
    facility_name : str
        正規化する施設名
    client : OpenAI, optional
        OpenAIクライアント（Noneの場合はAPIキーを環境変数から取得）
    cache : dict, optional
        キャッシュ辞書（同じ名前の再処理を避ける）
    
    Returns:
    --------
    str
        正規化された施設名（沖縄県内の施設の場合は標準名、それ以外は元の名前）
    """
    if pd.isna(facility_name) or not facility_name or str(facility_name).strip() == '':
        return ''
    
    facility_str = str(facility_name).strip()
    
    # キャッシュをチェック
    if cache is not None and facility_str in cache:
        return cache[facility_str]
    
    # クライアントが提供されていない場合は作成
    if client is None:
        api_key = os.getenv('OPENAI_API_KEY')
        if not api_key:
            print("[WARNING] OPENAI_API_KEY環境変数が設定されていません。施設名の正規化をスキップします。")
            return facility_str
        client = OpenAI(api_key=api_key)
    
    try:
        # OpenAI APIで施設名を正規化
        prompt = f"""以下の施設名が沖縄県内の医療施設かどうかを判定し、沖縄県内の施設の場合は標準名に正規化してください。
沖縄県内の主要な施設名の例：
- 県立宮古病院
- 県立北部病院
- 県立八重山病院
- 県立中部病院
- 琉大病院（琉球大学医学部附属病院）
- 南部医療センター
- 宮古病院
- 北部病院
- 中部病院
- 八重山病院
- 各種診療所（伊平屋、伊是名、西表、小浜、座間味、阿嘉、大原、粟国、渡名喜、波照間、北大東、南大東など）

入力施設名: {facility_str}

JSON形式で回答してください：
{{"is_okinawa": true/false, "normalized_name": "正規化された施設名（沖縄県内の場合のみ）", "original": "{facility_str}"}}
沖縄県外の施設の場合は、is_okinawaをfalseにし、normalized_nameは空文字列にしてください。"""
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "あなたは医療施設名を正規化する専門家です。JSON形式で正確に回答してください。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1
        )
        
        result_text = response.choices[0].message.content.strip()
        
        # JSONをパース
        # コードブロックがある場合は除去
        if "```json" in result_text:
            result_text = result_text.split("```json")[1].split("```")[0].strip()
        elif "```" in result_text:
            result_text = result_text.split("```")[1].split("```")[0].strip()
        
        result = json.loads(result_text)
        
        if result.get("is_okinawa", False) and result.get("normalized_name"):
            normalized = result["normalized_name"]
        else:
            normalized = facility_str
        
        # キャッシュに保存
        if cache is not None:
            cache[facility_str] = normalized
        
        return normalized
        
    except Exception as e:
        print(f"[WARNING] 施設名の正規化でエラーが発生しました ({facility_str}): {e}")
        # エラー時は元の名前を返す
        if cache is not None:
            cache[facility_str] = facility_str
        return facility_str


def is_okinawa_birthplace(birthplace: str) -> bool:
    """
    本籍が沖縄県かどうかを判定する
    
    Parameters:
    -----------
    birthplace : str
        本籍地
    
    Returns:
    --------
    bool
        沖縄県出身の場合True
    """
    if pd.isna(birthplace) or not birthplace:
        return False
    
    birthplace_str = str(birthplace).strip()
    
    # 沖縄県に関連するキーワードをチェック
    okinawa_keywords = ['沖縄', 'おきなわ', 'OKINAWA', 'okinawa']
    
    return any(keyword in birthplace_str for keyword in okinawa_keywords)


def is_okinawa_facility(facility_name: str, normalized_facilities: Optional[set] = None) -> bool:
    """
    施設が沖縄県内の施設かどうかを判定する
    
    Parameters:
    -----------
    facility_name : str
        施設名
    normalized_facilities : set, optional
        正規化済みの沖縄県内施設名のセット
    
    Returns:
    --------
    bool
        沖縄県内の施設の場合True
    """
    if pd.isna(facility_name) or not facility_name:
        return False
    
    facility_str = str(facility_name).strip()
    
    # 「次年度」を含む施設名は県外扱い（まだ確定していないため）
    # 例: "次年度中部病院", "次年度 中部病院" など
    if '次年度' in facility_str:
        return False
    
    # 基本的なキーワードチェック（常に実行）
    okinawa_keywords = [
        '県立', '宮古', '北部', '八重山', '中部', '南部', '琉大', '琉球大学',
        '伊平屋', '伊是名', '西表', '小浜', '座間味', '阿嘉', '大原', '粟国',
        '渡名喜', '波照間', '北大東', '南大東', '精和', '浦添', '沖縄'
    ]
    
    # キーワードチェック
    if any(keyword in facility_str for keyword in okinawa_keywords):
        return True
    
    # 正規化済みセットが提供されている場合はそれもチェック
    if normalized_facilities is not None:
        if facility_str in normalized_facilities:
            return True
    
    return False


def classify_status(status: str) -> str:
    """
    進路を分類する
    
    Parameters:
    -----------
    status : str
        進路
    
    Returns:
    --------
    str
        分類結果（'研修中', '転出', '修了', '中断', '退職'）
    """
    if pd.isna(status) or not status:
        return '研修中'
    
    status_str = str(status).strip()
    
    if status_str in ['転出', '修了']:
        return status_str
    elif status_str == '中断':
        return '中断'
    elif status_str == '退職':
        return '退職'
    else:
        # 進級などその他は研修中として扱う
        return '研修中'


def get_names_by_category(df_final: pd.DataFrame, facility_cache: Dict[str, str], 
                         okinawa_facilities_set: set, client: Optional[OpenAI] = None) -> Dict[str, list]:
    """
    各カテゴリに該当する名前のリストを取得する
    
    Parameters:
    -----------
    df_final : pd.DataFrame
        最終記録のDataFrame
    facility_cache : dict
        施設名の正規化キャッシュ
    okinawa_facilities_set : set
        沖縄県内施設名のセット
    client : OpenAI, optional
        OpenAIクライアント
    
    Returns:
    --------
    dict
        カテゴリごとの名前リストの辞書
    """
    names_by_category = {
        '研修中': [],
        '沖縄出身_沖縄内_転出・修了': [],
        '沖縄出身_沖縄外_転出・修了': [],
        '沖縄外出身_沖縄内_転出・修了': [],
        '沖縄外出身_沖縄外_転出・修了': [],
        '中断': [],
        '退職': []
    }
    
    # 本籍と動向調査の列を確認
    birthplace_col = '本籍' if '本籍' in df_final.columns else None
    facility_col = '動向調査' if '動向調査' in df_final.columns else None
    status_col = '進路' if '進路' in df_final.columns else None
    name_col = '名前' if '名前' in df_final.columns else None
    
    if not status_col or not name_col:
        return names_by_category
    
    for _, row in df_final.iterrows():
        status = classify_status(row[status_col] if status_col else '')
        name = row[name_col] if name_col and name_col in row.index else ''
        
        if pd.isna(name) or not name:
            continue
        
        if status == '研修中':
            names_by_category['研修中'].append(str(name))
        elif status == '中断':
            names_by_category['中断'].append(str(name))
        elif status == '退職':
            names_by_category['退職'].append(str(name))
        elif status in ['転出', '修了']:
            # 出身地を判定
            birthplace = row[birthplace_col] if birthplace_col and birthplace_col in row.index else ''
            is_okinawa_born = is_okinawa_birthplace(birthplace)
            
            # 転出先施設を取得
            facility = row[facility_col] if facility_col and facility_col in row.index else ''
            
            # 施設名を正規化
            normalized_facility = normalize_facility_name(facility, client=client, cache=facility_cache) if facility else ''
            
            # 沖縄県内施設かどうかを判定
            is_okinawa_facility_flag = is_okinawa_facility(normalized_facility, okinawa_facilities_set) if normalized_facility else False
            
            # カテゴリを決定（転出と修了は一緒にカウント）
            if is_okinawa_born:
                if is_okinawa_facility_flag:
                    key = '沖縄出身_沖縄内_転出・修了'
                else:
                    key = '沖縄出身_沖縄外_転出・修了'
            else:
                if is_okinawa_facility_flag:
                    key = '沖縄外出身_沖縄内_転出・修了'
                else:
                    key = '沖縄外出身_沖縄外_転出・修了'
            
            names_by_category[key].append(str(name))
    
    return names_by_category


def calculate_statistics(df_final: pd.DataFrame, facility_cache: Dict[str, str], 
                        okinawa_facilities_set: set, client: Optional[OpenAI] = None) -> Dict[str, int]:
    """
    各期の統計を計算する
    
    Parameters:
    -----------
    df_final : pd.DataFrame
        最終記録のDataFrame
    facility_cache : dict
        施設名の正規化キャッシュ
    okinawa_facilities_set : set
        沖縄県内施設名のセット
    
    Returns:
    --------
    dict
        統計結果の辞書
    """
    stats = {
        '研修中': 0,
        '沖縄出身_沖縄内_転出・修了': 0,
        '沖縄出身_沖縄外_転出・修了': 0,
        '沖縄外出身_沖縄内_転出・修了': 0,
        '沖縄外出身_沖縄外_転出・修了': 0,
        '中断': 0,
        '退職': 0
    }
    
    # 本籍と動向調査の列を確認
    birthplace_col = '本籍' if '本籍' in df_final.columns else None
    facility_col = '動向調査' if '動向調査' in df_final.columns else None
    status_col = '進路' if '進路' in df_final.columns else None
    
    if not status_col:
        return stats
    
    for _, row in df_final.iterrows():
        status = classify_status(row[status_col] if status_col else '')
        
        if status == '研修中':
            stats['研修中'] += 1
        elif status == '中断':
            stats['中断'] += 1
        elif status == '退職':
            stats['退職'] += 1
        elif status in ['転出', '修了']:
            # 出身地を判定
            birthplace = row[birthplace_col] if birthplace_col and birthplace_col in row.index else ''
            is_okinawa_born = is_okinawa_birthplace(birthplace)
            
            # 転出先施設を取得
            facility = row[facility_col] if facility_col and facility_col in row.index else ''
            
            # 施設名を正規化
            normalized_facility = normalize_facility_name(facility, client=client, cache=facility_cache) if facility else ''
            
            # 沖縄県内施設かどうかを判定
            is_okinawa_facility_flag = is_okinawa_facility(normalized_facility, okinawa_facilities_set) if normalized_facility else False
            
            # カテゴリを決定（転出と修了は一緒にカウント）
            if is_okinawa_born:
                if is_okinawa_facility_flag:
                    key = '沖縄出身_沖縄内_転出・修了'
                else:
                    key = '沖縄出身_沖縄外_転出・修了'
            else:
                if is_okinawa_facility_flag:
                    key = '沖縄外出身_沖縄内_転出・修了'
                else:
                    key = '沖縄外出身_沖縄外_転出・修了'
            
            stats[key] = stats.get(key, 0) + 1
    
    return stats


def create_period_sheets_from_master(master_file="研修医マスタ.xlsm", output_file="研修医データ_期別.xlsx"):
    """
    研修医マスタ.xlsmファイルのmainシートから各期ごとのシートを作成する
    括弧付き名前は同一人物として扱う

    Parameters:
    -----------
    master_file : str
        入力ファイル名またはパス（デフォルト: 研修医マスタ.xlsm）
    output_file : str
        出力Excelファイル名またはパス（デフォルト: 研修医データ_期別.xlsx）
    """

    # 絶対パスの場合はそのまま使用、相対パスの場合はカレントディレクトリから取得
    if os.path.isabs(master_file):
        master_path = master_file
    else:
        current_dir = os.getcwd()
        master_path = os.path.join(current_dir, master_file)
    
    if os.path.isabs(output_file):
        output_path = output_file
    else:
        current_dir = os.getcwd()
        output_path = os.path.join(current_dir, output_file)

    # ファイルの存在確認
    if not os.path.exists(master_path):
        raise FileNotFoundError(f"ファイルが見つかりません: {master_path}")

    print(f"マスターファイル: {master_path}")
    print(f"出力ファイル: {output_path}")

    # mainシートからマスターデータを読み込む（ヘッダー行を探す）
    print("\n研修医マスターデータ（mainシート）を読み込み中...")

    # まず最初の数行を確認してヘッダー行を探す
    df_temp = pd.read_excel(master_path, sheet_name='main', engine='openpyxl', header=None, nrows=10)
    print("\n最初の10行を確認:")
    print(df_temp)

    # ヘッダー行を探す（'年度'や'名前'などの列が含まれる行）
    header_row = None
    for idx in range(len(df_temp)):
        row_values = df_temp.iloc[idx].values
        row_str = ' '.join([str(v) for v in row_values if pd.notna(v)])
        if '年度' in row_str and '名前' in row_str:
            header_row = idx
            print(f"\n[OK] ヘッダー行を発見: {idx + 1}行目")
            print(f"ヘッダー内容: {row_values}")
            break

    if header_row is None:
        raise ValueError("ヘッダー行が見つかりません。'年度'と'名前'を含む行を探しましたが見つかりませんでした。")

    # ヘッダー行を指定して再度読み込み
    df_master = pd.read_excel(master_path, sheet_name='main', engine='openpyxl', header=header_row)

    # 空白列を削除（Unnamed列や完全に空の列）
    df_master = df_master.loc[:, ~df_master.columns.str.contains('^Unnamed')]
    df_master = df_master.dropna(axis=1, how='all')

    # 空白行を削除
    df_master = df_master.dropna(how='all')

    print(f"\n読み込み完了: {len(df_master)}行")
    print(f"列名: {list(df_master.columns)}")

    # 正規化された名前列を追加
    df_master['名前_正規化'] = df_master['名前'].apply(normalize_name)

    # 括弧付き名前の例を表示
    bracket_names = df_master[df_master['名前'].astype(str).str.contains(r'[（(]', na=False)]
    if len(bracket_names) > 0:
        print(f"\n[INFO] 括弧付き名前が見つかりました（{len(bracket_names)}件）:")
        for idx, row in bracket_names.head(10).iterrows():
            print(f"  元の名前: '{row['名前']}' -> 正規化後: '{row['名前_正規化']}'")
        if len(bracket_names) > 10:
            print(f"  ... 他 {len(bracket_names) - 10}件")

    # 必須列の確認
    required_columns = ['年度', '学年', '名前', '進路']
    missing_columns = [col for col in required_columns if col not in df_master.columns]

    if missing_columns:
        print(f"\n[WARNING] 警告: 必須列が見つかりません: {missing_columns}")
        print("利用可能な列:", list(df_master.columns))
        raise ValueError(f"必須列が不足しています: {missing_columns}")

    # '初・後'列の確認（異なる名前の可能性を考慮）
    ki_column = None
    for col in df_master.columns:
        if '初' in str(col) or '後' in str(col) or '期' in str(col):
            ki_column = col
            print(f"[OK] 期を表す列を発見: '{ki_column}'")
            break

    if ki_column is None:
        raise ValueError("'初・後'または期を表す列が見つかりません")

    # 終了進路のリスト
    end_statuses = ['転出', '修了', '中断', '退職']

    # 期番号を抽出（47期〜59期）
    df_master['期番号'] = df_master[ki_column].astype(str).str.extract(r'(\d+)期')[0].astype(float)
    df_filtered = df_master[df_master['期番号'].between(47, 59)]

    print(f"\n総データ数: {len(df_master)}行")
    print(f"47期〜59期のデータ数: {len(df_filtered)}行")

    if len(df_filtered) == 0:
        print("[WARNING] 警告: 47期〜59期のデータが見つかりません")
        print(f"期番号の内訳:")
        print(df_master['期番号'].value_counts().sort_index())
        raise ValueError("処理対象のデータがありません")

    # 新しいExcelファイルを作成
    wb = Workbook()
    # デフォルトシートを削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # マスターシートを作成（元データ全体）
    ws_master = wb.create_sheet('研修医マスター', 0)

    # マスターシートにヘッダーを書き込み（名前_正規化は除く）
    headers = [col for col in df_master.columns if col not in ['名前_正規化', '期番号']]
    for col_num, header in enumerate(headers, 1):
        cell = ws_master.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(name='Aptos Narrow', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # マスターシートにデータを書き込み
    for row_num, (_, row_data) in enumerate(df_master.iterrows(), 2):
        for col_num, col_name in enumerate(headers, 1):
            cell = ws_master.cell(row=row_num, column=col_num)
            value = row_data[col_name]
            cell.value = value if pd.notna(value) else ''

    # マスターシートの列幅を自動調整
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws_master[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_master.column_dimensions[column_letter].width = adjusted_width

    print("\n[OK] 研修医マスターシートを作成しました")

    # 出力用の列名を設定
    output_columns = ['年度', '学年', ki_column, 'PHS', '名前', 'ふりがな', '性別',
                      '専門科', '進路', '動向調査', '本籍', '出身大学', '備考']

    # 存在しない列は除外
    output_columns = [col for col in output_columns if col in df_master.columns]

    # OpenAI APIクライアントの初期化（施設名正規化用）
    api_key = os.getenv('OPENAI_API_KEY')
    client = None
    if api_key:
        try:
            client = OpenAI(api_key=api_key)
            print("\n[OK] OpenAI APIクライアントを初期化しました")
        except Exception as e:
            print(f"[WARNING] OpenAI APIクライアントの初期化に失敗しました: {e}")
    else:
        print("[WARNING] OPENAI_API_KEY環境変数が設定されていません。施設名の正規化をスキップします。")

    # 施設名正規化のキャッシュ
    facility_cache = {}
    
    # 沖縄県内施設のセットを作成（正規化後の名前を収集）
    print("\n施設名の正規化を実行中...")
    okinawa_facilities_set = set()
    for facility in OKINAWA_FACILITIES_RAW:
        normalized = normalize_facility_name(facility, client=client, cache=facility_cache)
        if normalized:
            okinawa_facilities_set.add(normalized)
    
    print(f"[OK] 沖縄県内施設の正規化完了: {len(okinawa_facilities_set)}施設")
    
    # 集計結果を保存するリスト
    all_statistics = []

    # 各期ごとに処理
    for ki_num in range(47, 60):
        sheet_name = f'{ki_num}期'
        print(f"\n{sheet_name}を処理中...")

        # その期のデータを抽出
        df_ki = df_filtered[df_filtered['期番号'] == ki_num].copy()

        if len(df_ki) == 0:
            print(f"  {sheet_name}: データなし、スキップします")
            continue

        print(f"  全記録数: {len(df_ki)}行")

        # 正規化された名前でグループ化し、各人の最終年データを取得
        final_records = []

        for normalized_name, group in df_ki.groupby('名前_正規化'):
            if not normalized_name:  # 空の名前はスキップ
                continue

            # 終了進路の記録があるか確認
            end_records = group[group['進路'].isin(end_statuses)]

            if len(end_records) > 0:
                # 終了進路の中で年度が最大のものを選択
                final_record = end_records.loc[end_records['年度'].idxmax()]
            else:
                # 年度が最大のものを選択
                final_record = group.loc[group['年度'].idxmax()]

            final_records.append(final_record)

            # 同一人物で複数の名前がある場合は表示
            unique_names = group['名前'].unique()
            if len(unique_names) > 1:
                print(f"  [INFO] 同一人物として統合: {list(unique_names)} -> '{normalized_name}'")

        # DataFrameに変換
        df_final = pd.DataFrame(final_records)

        # ふりがなでソート（空白の場合は名前でソート）
        if 'ふりがな' in df_final.columns:
            df_final['ソートキー'] = df_final['ふりがな'].fillna(df_final['名前'])
            df_final = df_final.sort_values('ソートキー')
            df_final = df_final.drop(columns=['ソートキー'])
        else:
            df_final = df_final.sort_values('名前')

        # 不要な列を削除
        cols_to_drop = ['期番号', '名前_正規化']
        df_final = df_final.drop(columns=[col for col in cols_to_drop if col in df_final.columns])

        print(f"  [OK] 最終記録数: {len(df_final)}名（重複なし・括弧名統合済み）")

        # 新しいシートを作成
        ws = wb.create_sheet(sheet_name)

        # ヘッダー行を書き込み
        display_headers = ['年度', '学年', '初・後', 'PHS', '名前', 'ふりがな', '性別',
                           '専門科', '進路', '動向調査', '本籍', '出身大学', '備考']

        for col_num, header in enumerate(display_headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            # ヘッダースタイル設定
            cell.font = Font(name='Aptos Narrow', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # データ行を書き込み
        for row_num, (_, row_data) in enumerate(df_final.iterrows(), 2):
            for col_num, header in enumerate(display_headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                # '初・後'の場合はki_columnから取得
                col_name = ki_column if header == '初・後' else header
                if col_name in row_data.index:
                    value = row_data[col_name]
                    cell.value = value if pd.notna(value) else ''
                else:
                    cell.value = ''

        # 列幅を自動調整
        for col_num in range(1, len(display_headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0

            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 集計を実行
        statistics = calculate_statistics(df_final, facility_cache, okinawa_facilities_set, client=client)
        statistics['期'] = ki_num
        all_statistics.append(statistics)
        print(f"  [OK] 集計完了: 研修中={statistics['研修中']}名, 中断={statistics['中断']}名, 退職={statistics['退職']}名")

        # 各カテゴリの名前リストを取得
        names_by_category = get_names_by_category(df_final, facility_cache, okinawa_facilities_set, client=client)

        # 各期のシートに集計結果を追加
        last_data_row = len(df_final) + 1  # データの最後の行番号（ヘッダー含む）
        summary_start_row = last_data_row + 3  # データの下に2行空けて開始
        
        # 集計結果のタイトル
        title_cell = ws.cell(row=summary_start_row - 1, column=1)
        title_cell.value = f"{sheet_name} 集計結果"
        title_cell.font = Font(name='Aptos Narrow', size=12, bold=True)
        title_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # 集計結果のヘッダー（3列: カテゴリ、人数、名前リスト）
        summary_headers = ['カテゴリ', '人数', '名前リスト']
        
        # 集計結果のデータ
        summary_data = [
            ['研修中', statistics['研修中'], names_by_category['研修中']],
            ['沖縄出身 → 沖縄内（転出・修了）', statistics['沖縄出身_沖縄内_転出・修了'], names_by_category['沖縄出身_沖縄内_転出・修了']],
            ['沖縄出身 → 沖縄外（転出・修了）', statistics['沖縄出身_沖縄外_転出・修了'], names_by_category['沖縄出身_沖縄外_転出・修了']],
            ['沖縄外出身 → 沖縄内（転出・修了）', statistics['沖縄外出身_沖縄内_転出・修了'], names_by_category['沖縄外出身_沖縄内_転出・修了']],
            ['沖縄外出身 → 沖縄外（転出・修了）', statistics['沖縄外出身_沖縄外_転出・修了'], names_by_category['沖縄外出身_沖縄外_転出・修了']],
            ['中断', statistics['中断'], names_by_category['中断']],
            ['退職', statistics['退職'], names_by_category['退職']]
        ]
        
        # ヘッダー行を書き込み
        for col_num, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=summary_start_row, column=col_num)
            cell.value = header
            cell.font = Font(name='Aptos Narrow', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # データ行を書き込み
        current_row = summary_start_row + 1
        for row_idx, (category, count, names_list) in enumerate(summary_data):
            # カテゴリ列
            cell_category = ws.cell(row=current_row, column=1)
            cell_category.value = category
            cell_category.alignment = Alignment(horizontal='left', vertical='top')
            
            # 人数列
            cell_count = ws.cell(row=current_row, column=2)
            cell_count.value = count
            cell_count.alignment = Alignment(horizontal='center', vertical='top')
            
            # 名前リスト列（複数行にまたがる可能性がある）
            names_text = '、'.join(names_list) if names_list else ''
            cell_names = ws.cell(row=current_row, column=3)
            cell_names.value = names_text
            cell_names.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # 行の背景色を交互に設定（見やすくするため）
            if row_idx % 2 == 0:
                fill_color = 'F2F2F2'
            else:
                fill_color = 'FFFFFF'
            
            for col_num in range(1, 4):
                cell = ws.cell(row=current_row, column=col_num)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            # 名前リストが長い場合は行の高さを調整
            if names_text:
                # 名前の数に応じて行の高さを調整（1名あたり約15ピクセル、最大200ピクセル）
                estimated_height = min(len(names_list) * 15 + 10, 200)
                ws.row_dimensions[current_row].height = estimated_height
            
            current_row += 1
        
        # 合計行を追加
        total_count = sum([
            statistics['研修中'],
            statistics['沖縄出身_沖縄内_転出・修了'],
            statistics['沖縄出身_沖縄外_転出・修了'],
            statistics['沖縄外出身_沖縄内_転出・修了'],
            statistics['沖縄外出身_沖縄外_転出・修了'],
            statistics['中断'],
            statistics['退職']
        ])
        
        # 合計行のセル
        cell_total_category = ws.cell(row=current_row, column=1)
        cell_total_category.value = '合計'
        cell_total_category.font = Font(name='Aptos Narrow', size=11, bold=True)
        cell_total_category.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cell_total_category.alignment = Alignment(horizontal='left', vertical='center')
        
        cell_total_count = ws.cell(row=current_row, column=2)
        cell_total_count.value = total_count
        cell_total_count.font = Font(name='Aptos Narrow', size=11, bold=True)
        cell_total_count.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cell_total_count.alignment = Alignment(horizontal='center', vertical='center')
        
        cell_total_names = ws.cell(row=current_row, column=3)
        cell_total_names.value = ''  # 合計行には名前リストは不要
        cell_total_names.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        
        # 集計結果の列幅を調整
        ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width, 35)
        ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width, 12)
        ws.column_dimensions['C'].width = max(ws.column_dimensions['C'].width, 50)  # 名前リスト用に広めに設定

    # 集計結果シートを作成
    print("\n集計結果シートを作成中...")
    ws_stats = wb.create_sheet('集計結果', 1)
    
    # ヘッダー行
    headers_stats = [
        '期', '総数', '研修中', 
        '沖縄出身_沖縄内_転出・修了',
        '沖縄出身_沖縄外_転出・修了',
        '沖縄外出身_沖縄内_転出・修了',
        '沖縄外出身_沖縄外_転出・修了',
        '中断', '退職'
    ]
    
    for col_num, header in enumerate(headers_stats, 1):
        cell = ws_stats.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(name='Aptos Narrow', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # データ行を書き込み
    for row_num, stats in enumerate(all_statistics, 2):
        # 各期の総数を計算
        total_count = sum([
            stats.get('研修中', 0),
            stats.get('沖縄出身_沖縄内_転出・修了', 0),
            stats.get('沖縄出身_沖縄外_転出・修了', 0),
            stats.get('沖縄外出身_沖縄内_転出・修了', 0),
            stats.get('沖縄外出身_沖縄外_転出・修了', 0),
            stats.get('中断', 0),
            stats.get('退職', 0)
        ])
        
        for col_num, header in enumerate(headers_stats, 1):
            cell = ws_stats.cell(row=row_num, column=col_num)
            if header == '総数':
                value = total_count
            else:
                value = stats.get(header, 0)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 合計行を追加
    total_row = len(all_statistics) + 2
    ws_stats.cell(row=total_row, column=1).value = '合計'
    ws_stats.cell(row=total_row, column=1).font = Font(bold=True)
    ws_stats.cell(row=total_row, column=1).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    for col_num in range(2, len(headers_stats) + 1):
        col_letter = get_column_letter(col_num)
        formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        cell = ws_stats.cell(row=total_row, column=col_num)
        cell.value = formula
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 列幅を自動調整
    for col_num in range(1, len(headers_stats) + 1):
        column_letter = get_column_letter(col_num)
        max_length = len(str(headers_stats[col_num - 1]))
        for row in ws_stats[column_letter]:
            try:
                if row.value and len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_stats.column_dimensions[column_letter].width = adjusted_width
    
    print("[OK] 集計結果シートを作成しました")

    # ファイルを保存
    print(f"\n\nファイルを保存中: {output_path}")
    wb.save(output_path)
    print("[OK] 完了！")

    # 各期の統計情報を表示
    print("\n" + "=" * 70)
    print("各期の研修医数（最終記録のみ・括弧名統合済み）")
    print("=" * 70)

    for ki_num in range(47, 60):
        df_ki = df_filtered[df_filtered['期番号'] == ki_num]
        if len(df_ki) > 0:
            # 正規化された名前でユニークカウント
            unique_count = df_ki['名前_正規化'].nunique()
            print(f"{ki_num}期: {unique_count}名")

    # 集計結果の詳細を表示
    print("\n" + "=" * 70)
    print("集計結果の詳細")
    print("=" * 70)
    
    for stats in all_statistics:
        ki = stats['期']
        total_count = sum([
            stats['研修中'],
            stats['沖縄出身_沖縄内_転出・修了'],
            stats['沖縄出身_沖縄外_転出・修了'],
            stats['沖縄外出身_沖縄内_転出・修了'],
            stats['沖縄外出身_沖縄外_転出・修了'],
            stats['中断'],
            stats['退職']
        ])
        print(f"\n{ki}期: 総数 {total_count}名")
        print(f"  研修中: {stats['研修中']}名")
        print(f"  沖縄出身 → 沖縄内（転出・修了）: {stats['沖縄出身_沖縄内_転出・修了']}名")
        print(f"  沖縄出身 → 沖縄外（転出・修了）: {stats['沖縄出身_沖縄外_転出・修了']}名")
        print(f"  沖縄外出身 → 沖縄内（転出・修了）: {stats['沖縄外出身_沖縄内_転出・修了']}名")
        print(f"  沖縄外出身 → 沖縄外（転出・修了）: {stats['沖縄外出身_沖縄外_転出・修了']}名")
        print(f"  中断: {stats['中断']}名")
        print(f"  退職: {stats['退職']}名")

    print("\n" + "=" * 70)
    print(f"[OK] 出力ファイル: {output_path}")
    print("=" * 70)


# 使用例
if __name__ == "__main__":
    try:
        # 基本的な使い方（デフォルト設定）
        create_period_sheets_from_master()
    except Exception as e:
        print(f"\n[ERROR] エラーが発生しました: {e}")
        import traceback

        traceback.print_exc()
