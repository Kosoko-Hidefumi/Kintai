"""
修了式資料生成スクリプト
- Sheet1のデータから進級者・転出者リストを作成
- 専門科で昇順ソート
- 中断・退職・病休者を除外
- 修了者は転出リストに含める
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


def load_data_from_sheet1(ws):
    """Sheet1からデータを読み込み、辞書のリストとして返す"""
    data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1] and row[4]:  # 学年と名前がある行のみ
            record = {
                '学年': row[1],  # B列
                '名前': row[4],  # E列
                '専門科': row[7] or '',  # H列
                '進路': row[8] or '',  # I列
                '動向調査': row[9] or '',  # J列
                '出身大学': row[11] or ''  # L列
            }
            data.append(record)
    return data


def classify_data(data):
    """データを進級者と転出者（修了者含む）に分類し、中断・退職・病休者を除外"""
    exclude_keywords = ['中断', '退職', '病休', '休職']

    shinkyu = []  # 進級者
    tenshutsu = []  # 転出者（修了者含む）

    for record in data:
        shinro = record['進路']

        # 除外対象をチェック
        if any(keyword in shinro for keyword in exclude_keywords):
            continue

        if '進級' in shinro:
            shinkyu.append(record)
        elif '転出' in shinro or '修了' in shinro:
            tenshutsu.append(record)

    return shinkyu, tenshutsu


def sort_by_specialty(records):
    """専門科で昇順ソート"""
    return sorted(records, key=lambda x: x['専門科'])


def group_by_grade(records):
    """学年ごとにグループ化"""
    groups = {}
    for record in records:
        grade = record['学年']
        if grade not in groups:
            groups[grade] = []
        groups[grade].append(record)
    return groups


def create_output_sheet(wb, shinkyu, tenshutsu):
    """修了式資料シートを作成"""
    sheet_name = '修了式資料_進級転出'

    # 既存シートがあれば削除
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # スタイル定義
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
    section_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        bottom=Side(style='thin', color='000000')
    )

    current_row = 1

    # タイトル
    ws.cell(row=current_row, column=1, value='修了式資料（進級・転出）')
    ws.cell(row=current_row, column=1).font = title_font
    current_row += 2

    # 【進級者】セクション
    ws.cell(row=current_row, column=1, value='【進級者】')
    ws.cell(row=current_row, column=1).font = section_font
    ws.cell(row=current_row, column=1).fill = section_fill
    current_row += 1

    # 進級者を学年別にグループ化してソート
    shinkyu_groups = group_by_grade(shinkyu)
    grade_order = ['PGY1', 'PGY2', 'PGY3', 'PGY4']

    for grade in grade_order:
        if grade not in shinkyu_groups:
            continue

        records = sort_by_specialty(shinkyu_groups[grade])

        # 学年ヘッダー
        ws.cell(row=current_row, column=1, value=f'{grade}（{len(records)}名）')
        ws.cell(row=current_row, column=1).font = header_font
        current_row += 1

        # 列ヘッダー
        headers = ['No.', '名前', '専門科', '出身大学']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        current_row += 1

        # データ行
        for idx, record in enumerate(records, 1):
            ws.cell(row=current_row, column=1, value=idx).alignment = center_align
            ws.cell(row=current_row, column=2, value=record['名前'])
            ws.cell(row=current_row, column=3, value=record['専門科'])
            ws.cell(row=current_row, column=4, value=record['出身大学'])
            current_row += 1

        current_row += 1  # 空行

    # 【転出者（修了者含む）】セクション
    ws.cell(row=current_row, column=1, value='【転出者（修了者含む）】')
    ws.cell(row=current_row, column=1).font = section_font
    ws.cell(row=current_row, column=1).fill = section_fill
    current_row += 1

    # 転出者を学年別にグループ化してソート
    tenshutsu_groups = group_by_grade(tenshutsu)
    tenshutsu_grade_order = ['PGY2', 'PGY4', 'PGY5']

    for grade in tenshutsu_grade_order:
        if grade not in tenshutsu_groups:
            continue

        records = sort_by_specialty(tenshutsu_groups[grade])

        # 学年ヘッダー
        ws.cell(row=current_row, column=1, value=f'{grade}（{len(records)}名）')
        ws.cell(row=current_row, column=1).font = header_font
        current_row += 1

        # 列ヘッダー
        headers = ['No.', '名前', '専門科', '転出先']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        current_row += 1

        # データ行
        for idx, record in enumerate(records, 1):
            ws.cell(row=current_row, column=1, value=idx).alignment = center_align
            ws.cell(row=current_row, column=2, value=record['名前'])
            ws.cell(row=current_row, column=3, value=record['専門科'])
            # 転出先は動向調査列から取得、なければ進路を表示
            destination = record['動向調査'] if record['動向調査'] else record['進路']
            ws.cell(row=current_row, column=4, value=destination)
            current_row += 1

        current_row += 1  # 空行

    # 【集計】セクション
    ws.cell(row=current_row, column=1, value='【集計】')
    ws.cell(row=current_row, column=1).font = section_font
    ws.cell(row=current_row, column=1).fill = section_fill
    current_row += 1

    ws.cell(row=current_row, column=1, value='進級者総数')
    ws.cell(row=current_row, column=2, value=len(shinkyu))
    current_row += 1

    ws.cell(row=current_row, column=1, value='転出・修了者総数')
    ws.cell(row=current_row, column=2, value=len(tenshutsu))

    # 列幅調整
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35

    return ws


def main():
    # 同階層のExcelファイルを読み込み（ファイル名を適宜変更）
    script_dir = Path(__file__).parent
    excel_files = list(script_dir.glob('*.xlsx'))

    if not excel_files:
        print('エラー: 同階層にExcelファイルが見つかりません')
        return

    # 最初に見つかったExcelファイルを使用（複数ある場合は指定が必要）
    file_path = excel_files[0]
    print(f'処理対象ファイル: {file_path.name}')

    # ワークブックを読み込み
    wb = openpyxl.load_workbook(file_path)

    # Sheet1からデータを取得
    if 'Sheet1' not in wb.sheetnames:
        print('エラー: Sheet1が見つかりません')
        return

    sheet1 = wb['Sheet1']

    # データ読み込み
    data = load_data_from_sheet1(sheet1)
    print(f'読み込んだデータ件数: {len(data)}件')

    # データ分類
    shinkyu, tenshutsu = classify_data(data)
    print(f'進級者: {len(shinkyu)}名')
    print(f'転出・修了者: {len(tenshutsu)}名')

    # 出力シート作成
    create_output_sheet(wb, shinkyu, tenshutsu)

    # 保存
    wb.save(file_path)
    print(f'\n修了式資料を生成しました: 修了式資料_進級転出')
    print(f'ファイル保存完了: {file_path.name}')


if __name__ == '__main__':
    main()
